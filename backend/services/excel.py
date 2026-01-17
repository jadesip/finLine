"""
finLine Excel Export Service

Generates Excel workbooks with LBO analysis results.
"""

import io
import logging
from datetime import datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


class ExcelExporter:
    """Exports LBO analysis to Excel workbook."""

    # Styles
    HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
    METRIC_FONT = Font(bold=True, size=10)
    NUMBER_FONT = Font(size=10)
    THIN_BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    PERCENT_FORMAT = "0.0%"
    NUMBER_FORMAT = "#,##0.0"
    CURRENCY_FORMAT = "$#,##0.0"

    def __init__(self):
        logger.info("ExcelExporter initialized")

    def export_analysis(
        self,
        project_data: dict[str, Any],
        analysis_result: dict[str, Any],
        case_id: str = "base_case",
    ) -> bytes:
        """
        Export LBO analysis to Excel.

        Args:
            project_data: Full project data
            analysis_result: Result from run_lbo_analysis()
            case_id: Which case to export

        Returns:
            Excel file as bytes
        """
        wb = Workbook()

        # Remove default sheet
        default_sheet = wb.active
        wb.remove(default_sheet)

        # Create sheets
        self._create_summary_sheet(wb, project_data, analysis_result, case_id)
        self._create_financials_sheet(wb, project_data, analysis_result, case_id)
        self._create_debt_sheet(wb, analysis_result)
        self._create_returns_sheet(wb, analysis_result)

        # Save to bytes
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        logger.info(f"Excel export completed for case {case_id}")
        return buffer.getvalue()

    def _create_summary_sheet(
        self,
        wb: Workbook,
        project_data: dict[str, Any],
        analysis: dict[str, Any],
        case_id: str,
    ):
        """Create summary sheet with key metrics."""
        ws = wb.create_sheet("Summary", 0)

        meta = project_data.get("meta", {})
        returns = analysis.get("returns", {})
        sources_uses = analysis.get("sources_and_uses", {})

        # Title
        ws["A1"] = f"LBO Analysis: {meta.get('company_name', meta.get('name', 'Project'))}"
        ws["A1"].font = Font(bold=True, size=16)
        ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws["A2"].font = Font(italic=True, size=10)

        # Key Returns
        row = 4
        ws[f"A{row}"] = "KEY RETURNS"
        ws[f"A{row}"].font = Font(bold=True, size=12)
        ws.merge_cells(f"A{row}:B{row}")

        row += 1
        metrics = [
            ("IRR", returns.get("irr"), self.PERCENT_FORMAT),
            ("MOIC", returns.get("moic"), "0.00x"),
            ("Entry Equity", returns.get("entry_equity"), self.CURRENCY_FORMAT),
            ("Exit Equity", returns.get("exit_equity"), self.CURRENCY_FORMAT),
            ("Holding Period (years)", returns.get("holding_period_years"), "0.0"),
        ]

        for name, value, fmt in metrics:
            ws[f"A{row}"] = name
            ws[f"A{row}"].font = self.METRIC_FONT
            ws[f"B{row}"] = value
            ws[f"B{row}"].number_format = fmt
            row += 1

        # Sources & Uses
        row += 1
        ws[f"A{row}"] = "SOURCES & USES"
        ws[f"A{row}"].font = Font(bold=True, size=12)
        ws.merge_cells(f"A{row}:B{row}")

        row += 1
        ws[f"A{row}"] = "Sources"
        ws[f"A{row}"].font = Font(bold=True, underline="single")
        row += 1

        sources = sources_uses.get("sources", {})
        for name, value in sources.items():
            ws[f"A{row}"] = name.replace("_", " ").title()
            ws[f"B{row}"] = value
            ws[f"B{row}"].number_format = self.CURRENCY_FORMAT
            row += 1

        row += 1
        ws[f"A{row}"] = "Uses"
        ws[f"A{row}"].font = Font(bold=True, underline="single")
        row += 1

        uses = sources_uses.get("uses", {})
        for name, value in uses.items():
            ws[f"A{row}"] = name.replace("_", " ").title()
            ws[f"B{row}"] = value
            ws[f"B{row}"].number_format = self.CURRENCY_FORMAT
            row += 1

        # Column widths
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 15

    def _create_financials_sheet(
        self,
        wb: Workbook,
        project_data: dict[str, Any],
        analysis: dict[str, Any],
        case_id: str,
    ):
        """Create financials sheet with P&L and cash flow."""
        ws = wb.create_sheet("Financials")

        case_data = project_data.get("cases", {}).get(case_id, {})
        financials = case_data.get("financials", {})
        cash_flows = analysis.get("cash_flows", [])

        # Get years from cash flows
        years = [str(cf.get("year", "")) for cf in cash_flows]

        # Header row
        row = 1
        ws["A1"] = "Income Statement"
        ws["A1"].font = Font(bold=True, size=12)

        row = 2
        ws[f"A{row}"] = "Metric"
        ws[f"A{row}"].font = self.HEADER_FONT
        ws[f"A{row}"].fill = self.HEADER_FILL

        for i, year in enumerate(years):
            col = get_column_letter(i + 2)
            ws[f"{col}{row}"] = year
            ws[f"{col}{row}"].font = self.HEADER_FONT
            ws[f"{col}{row}"].fill = self.HEADER_FILL
            ws[f"{col}{row}"].alignment = Alignment(horizontal="center")

        # Income statement metrics
        income_stmt = financials.get("income_statement", {})
        metrics = [
            ("Revenue", "revenue"),
            ("EBITDA", "ebitda"),
            ("D&A", "d_and_a"),
            ("EBIT", "ebit"),
        ]

        row = 3
        for label, key in metrics:
            ws[f"A{row}"] = label
            ws[f"A{row}"].font = self.METRIC_FONT

            data = income_stmt.get(key, [])
            data_dict = {str(d.get("year", "")): d.get("value", 0) for d in data if isinstance(d, dict)}

            for i, year in enumerate(years):
                col = get_column_letter(i + 2)
                value = data_dict.get(year, 0) or 0
                ws[f"{col}{row}"] = value
                ws[f"{col}{row}"].number_format = self.NUMBER_FORMAT
                ws[f"{col}{row}"].alignment = Alignment(horizontal="right")

            row += 1

        # Cash Flow section
        row += 1
        ws[f"A{row}"] = "Cash Flow"
        ws[f"A{row}"].font = Font(bold=True, size=12)

        row += 1
        cf_metrics = [
            ("CFADS", "cfads"),
            ("Capex", "capex"),
            ("ΔWC", "delta_wc"),
            ("Interest", "interest_expense"),
            ("Debt Repayment", "debt_repayment"),
            ("Cash Balance", "ending_cash"),
        ]

        for label, key in cf_metrics:
            ws[f"A{row}"] = label
            ws[f"A{row}"].font = self.METRIC_FONT

            for i, cf in enumerate(cash_flows):
                col = get_column_letter(i + 2)
                value = cf.get(key, 0) or 0
                ws[f"{col}{row}"] = value
                ws[f"{col}{row}"].number_format = self.NUMBER_FORMAT
                ws[f"{col}{row}"].alignment = Alignment(horizontal="right")

            row += 1

        # Column widths
        ws.column_dimensions["A"].width = 20
        for i in range(len(years)):
            ws.column_dimensions[get_column_letter(i + 2)].width = 12

    def _create_debt_sheet(self, wb: Workbook, analysis: dict[str, Any]):
        """Create debt schedule sheet."""
        ws = wb.create_sheet("Debt Schedule")

        debt_schedule = analysis.get("debt_schedule", [])
        if not debt_schedule:
            ws["A1"] = "No debt data available"
            return

        # Header
        row = 1
        headers = [
            "Year",
            "Opening Balance",
            "Drawdown",
            "Repayment",
            "PIK Interest",
            "Closing Balance",
            "Cash Interest",
            "Total Interest",
        ]

        for i, header in enumerate(headers):
            col = get_column_letter(i + 1)
            ws[f"{col}{row}"] = header
            ws[f"{col}{row}"].font = self.HEADER_FONT
            ws[f"{col}{row}"].fill = self.HEADER_FILL
            ws[f"{col}{row}"].alignment = Alignment(horizontal="center")

        # Data rows
        row = 2
        for period in debt_schedule:
            ws[f"A{row}"] = period.get("year", "")
            ws[f"B{row}"] = period.get("opening_balance", 0)
            ws[f"C{row}"] = period.get("drawdown", 0)
            ws[f"D{row}"] = period.get("repayment", 0)
            ws[f"E{row}"] = period.get("pik_interest", 0)
            ws[f"F{row}"] = period.get("closing_balance", 0)
            ws[f"G{row}"] = period.get("cash_interest", 0)
            ws[f"H{row}"] = period.get("total_interest", 0)

            for i in range(1, 9):
                col = get_column_letter(i)
                if i > 1:
                    ws[f"{col}{row}"].number_format = self.NUMBER_FORMAT
                ws[f"{col}{row}"].alignment = Alignment(horizontal="right" if i > 1 else "center")

            row += 1

        # Column widths
        widths = [8, 15, 12, 12, 12, 15, 12, 12]
        for i, width in enumerate(widths):
            ws.column_dimensions[get_column_letter(i + 1)].width = width

    def _create_returns_sheet(self, wb: Workbook, analysis: dict[str, Any]):
        """Create returns analysis sheet."""
        ws = wb.create_sheet("Returns")

        returns = analysis.get("returns", {})
        sources_uses = analysis.get("sources_and_uses", {})

        # Title
        ws["A1"] = "Returns Analysis"
        ws["A1"].font = Font(bold=True, size=14)

        # Key metrics
        row = 3
        ws[f"A{row}"] = "KEY METRICS"
        ws[f"A{row}"].font = Font(bold=True, size=12)
        ws.merge_cells(f"A{row}:C{row}")

        row += 1
        metrics = [
            ("IRR", returns.get("irr"), self.PERCENT_FORMAT, "Internal Rate of Return"),
            ("MOIC", returns.get("moic"), "0.00x", "Multiple on Invested Capital"),
            ("Entry Equity", returns.get("entry_equity"), self.CURRENCY_FORMAT, "Sponsor equity at close"),
            ("Exit Equity", returns.get("exit_equity"), self.CURRENCY_FORMAT, "Sponsor proceeds at exit"),
            ("Holding Period", returns.get("holding_period_years"), "0.0 years", "Investment duration"),
        ]

        for name, value, fmt, desc in metrics:
            ws[f"A{row}"] = name
            ws[f"A{row}"].font = self.METRIC_FONT
            ws[f"B{row}"] = value
            ws[f"B{row}"].number_format = fmt.replace(" years", "")
            ws[f"C{row}"] = desc
            ws[f"C{row}"].font = Font(italic=True, size=9, color="666666")
            row += 1

        # Valuation
        row += 1
        ws[f"A{row}"] = "VALUATION"
        ws[f"A{row}"].font = Font(bold=True, size=12)

        row += 1
        valuation_metrics = [
            ("Entry Enterprise Value", returns.get("entry_ev"), self.CURRENCY_FORMAT),
            ("Entry Multiple", returns.get("entry_multiple"), "0.0x"),
            ("Exit Enterprise Value", returns.get("exit_ev"), self.CURRENCY_FORMAT),
            ("Exit Multiple", returns.get("exit_multiple"), "0.0x"),
        ]

        for name, value, fmt in valuation_metrics:
            ws[f"A{row}"] = name
            ws[f"A{row}"].font = self.METRIC_FONT
            ws[f"B{row}"] = value
            ws[f"B{row}"].number_format = fmt
            row += 1

        # Column widths
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 30


def export_project_to_excel(
    project_data: dict[str, Any],
    analysis_result: dict[str, Any],
    case_id: str = "base_case",
) -> bytes:
    """
    Export project analysis to Excel.

    Convenience function wrapping ExcelExporter.
    """
    exporter = ExcelExporter()
    return exporter.export_analysis(project_data, analysis_result, case_id)
